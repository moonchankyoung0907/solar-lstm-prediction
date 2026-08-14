import os
import openpyxl

TXT_PATH = r"C:\Users\sejae\Desktop\웨더게이트 소프트웨어\WH-2300S PC소프트웨어 25.03 (유저용)\WH-2300S 관련 User 제공자료\WH-2300S display software\WH24Data.txt"
XLSX_PATH = r"C:\Users\sejae\Desktop\sensor_log.xlsx"

def parse_line(line):
    try:
        parts = line.strip().split(",")
        if len(parts) < 10:
            return None
        dt = parts[0].strip()
        # "▲"는 cp949 디코딩 과정에서 "°"가 깨진 것 (온도의 "▲C"와 동일한 현상). 풍향은 0~360도.
        wind_dir = parts[1].strip().replace("▲", "").strip()
        temp = parts[2].strip().replace("▲C","").replace("℃","").strip()
        humidity = parts[3].strip().replace("%","").strip()
        wind = parts[4].strip().replace("m/s","").strip()
        rain = parts[6].strip()
        uv = parts[7].strip()      # 자외선 원시값 (WH24 UART Display 화면의 "UV")
        uvi = parts[8].strip()     # 자외선 지수 0~11 (WH24 UART Display 화면의 "UVI")
        lux = parts[9].strip().replace("lux","").strip()
        # CRC 에러 체크
        crc = "OK"
        for part in parts:
            if "CRC" in part or "ERROR" in part:
                crc = "ERROR"
                break
        return [dt, float(temp), float(humidity), float(wind), float(rain), float(lux), crc, float(wind_dir), float(uv), float(uvi)]
    except:
        return None

def apply_interpolation(ws):
    max_row = ws.max_row
    data_cols = [2, 3, 4, 5, 6, 8, 9, 10]  # temperature, humidity, wind_speed, rainfall, light_lux, wind_direction, uv, uvi (7=crc_status 제외)
    for row_idx in range(2, max_row + 1):
        crc = ws.cell(row=row_idx, column=7).value
        if crc == "ERROR":
            for col in data_cols:
                prev_val = None
                next_val = None
                if row_idx > 2:
                    prev_val = ws.cell(row=row_idx-1, column=col).value
                if row_idx < max_row:
                    next_val = ws.cell(row=row_idx+1, column=col).value
                if prev_val is not None and next_val is not None:
                    ws.cell(row=row_idx, column=col).value = (prev_val + next_val) / 2
                elif prev_val is not None:
                    ws.cell(row=row_idx, column=col).value = prev_val
                elif next_val is not None:
                    ws.cell(row=row_idx, column=col).value = next_val

def main():
    if os.path.exists(XLSX_PATH):
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active
        header = [c.value for c in ws[1]]
        # 7번째 컬럼(crc_status) 라벨이 예전부터 비어있는 파일이 있어 데이터는 정상이어도
        # 헤더만 None인 경우가 있음 - 라벨만 보정
        if len(header) >= 7 and not header[6]:
            ws.cell(row=1, column=7, value="crc_status")
            header[6] = "crc_status"
        for col_name in ("wind_direction", "uv", "uvi"):
            if col_name not in header:
                ws.cell(row=1, column=len(header) + 1, value=col_name)
                header.append(col_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["datetime","temperature","humidity","wind_speed","rainfall","light_lux","crc_status","wind_direction","uv","uvi"])

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing.add(str(row[0]))

    new_count = 0
    with open(TXT_PATH, "r", encoding="cp949", errors="ignore") as f:
        for line in f:
            if line.strip() == "":
                continue
            row = parse_line(line)
            if row and str(row[0]) not in existing:
                ws.append(row)
                existing.add(str(row[0]))
                new_count += 1

    apply_interpolation(ws)
    wb.save(XLSX_PATH)
    print(f"완료! 새로 추가된 데이터: {new_count}개")

if __name__ == "__main__":
    main()